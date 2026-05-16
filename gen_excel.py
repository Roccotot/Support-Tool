#!/usr/bin/env python3
"""
Legge index.html, estrae i blocchi RAW e RAW_NOVPN,
e genera dispositivi_mancanti.xlsx con foglio "Dispositivi".
"""
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

HTML_PATH  = "/home/user/NOC-DASHBOARD-ROCCO/index.html"
EXCEL_PATH = "/home/user/NOC-DASHBOARD-ROCCO/dispositivi_mancanti.xlsx"

# ── 1. Leggi il file HTML ─────────────────────────────────────────────────────
with open(HTML_PATH, encoding="utf-8") as fh:
    html = fh.read()

# ── 2. Estrai i due template literal ─────────────────────────────────────────
def extract_template(html, var_name):
    """Ritorna il contenuto del template literal di 'var_name'."""
    pattern = rf"const\s+{re.escape(var_name)}\s*=\s*`(.*?)`"
    m = re.search(pattern, html, re.DOTALL)
    if not m:
        raise ValueError(f"Template literal '{var_name}' non trovato")
    return m.group(1)

raw_vpn   = extract_template(html, "RAW")
raw_novpn = extract_template(html, "RAW_NOVPN")

# ── 3. Placeholder riconosciuti ───────────────────────────────────────────────
PLACEHOLDERS = {"Proiettore", "Server", "Processore Audio", "Processore audio", "Processore"}

# ── 4. Parser delle righe ────────────────────────────────────────────────────
def parse_block(raw_text, tipo):
    """
    Ritorna una lista di dict con chiavi:
        cinema, city, tipo, sala, device, ip
    """
    rows = []
    for line in raw_text.split("\n"):
        # Sostituisci sequenze letterali \t con tab reale (già gestiti dal file,
        # ma per sicurezza lo normalizziamo)
        line = line.replace("\\t", "\t")
        line = line.strip()
        if not line:
            continue
        # Commenti JavaScript
        if line.startswith("//"):
            continue

        # Separa per TAB (reale)
        parts = line.split("\t")
        if len(parts) < 2:
            continue

        desc = parts[0].strip()
        ip   = parts[1].strip() if len(parts) > 1 else ""

        # Riga coordinate → ignora
        if "Coord" in desc and (len(parts) >= 3 and parts[2].strip() == "GEO"):
            continue
        # Anche se non c'è la terza colonna ma la seconda contiene lat,lon
        if re.match(r'^[\d.]+,[\d.]+$', ip):
            continue

        # Analizza la parte descrittiva: "Cinema - Città - Sala - Dispositivo"
        seg = [s.strip() for s in desc.split(" - ")]
        if len(seg) < 4:
            # Potrebbe mancare la sala (es. "Cinema - Città - Dispositivo")
            # ma in realtà non è un pattern standard → skip
            continue

        cinema = seg[0]
        city   = seg[1]
        sala   = " - ".join(seg[2:-1])   # tutto ciò che sta tra città e dispositivo
        device = seg[-1]

        rows.append({
            "cinema": cinema,
            "city":   city,
            "tipo":   tipo,
            "sala":   sala,
            "device": device,
            "ip":     ip,
        })
    return rows

entries = parse_block(raw_vpn,   "VPN") + parse_block(raw_novpn, "Offline")

# ── 5. Crea il workbook Excel ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Dispositivi"

# Colonne
COLS = ["Cinema", "Città", "Tipo", "Sala", "Dispositivo", "IP"]
ws.append(COLS)

# Stile intestazione
hdr_fill = PatternFill("solid", fgColor="1F3864")
hdr_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Colori per placeholder
fill_proj  = PatternFill("solid", fgColor="FFB3B3")   # Proiettore  → rosso chiaro
fill_serv  = PatternFill("solid", fgColor="FFFF99")   # Server      → giallo chiaro
fill_audio = PatternFill("solid", fgColor="FFD580")   # Proc. Audio → arancione chiaro

def placeholder_fill(device):
    d = device.strip()
    if d == "Proiettore":
        return fill_proj
    if d == "Server":
        return fill_serv
    if d in ("Processore Audio", "Processore audio", "Processore"):
        return fill_audio
    return None

# Scrivi le righe
for e in entries:
    row_vals = [e["cinema"], e["city"], e["tipo"], e["sala"], e["device"], e["ip"]]
    ws.append(row_vals)
    fill = placeholder_fill(e["device"])
    if fill:
        for cell in ws[ws.max_row]:
            cell.fill = fill

# Larghezze colonne (approssimative)
col_widths = [30, 28, 10, 25, 30, 22]
for i, w in enumerate(col_widths, start=1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Salva
wb.save(EXCEL_PATH)
print(f"Salvato: {EXCEL_PATH}")
print(f"Righe scritte (esclusa intestazione): {len(entries)}")
