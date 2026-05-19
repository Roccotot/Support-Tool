#!/usr/bin/env python3
"""
Genera dispositivi_mancanti.xlsx leggendo i template RAW e RAW_NOVPN da index.html.
"""

import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# ── Leggi index.html ──────────────────────────────────────────────────────────
with open('/home/user/NOC-DASHBOARD-ROCCO/index.html', 'r', encoding='utf-8') as f:
    html = f.read()

# ── Estrai i blocchi RAW e RAW_NOVPN ─────────────────────────────────────────
def extract_block(text, varname):
    """Estrae il contenuto del template literal `const VARNAME = \`...\`;`"""
    pattern = r'const\s+' + re.escape(varname) + r'\s*=\s*`(.*?)`\s*;'
    m = re.search(pattern, text, re.DOTALL)
    if not m:
        raise ValueError(f"Blocco {varname} non trovato in index.html")
    return m.group(1)

raw_vpn    = extract_block(html, 'RAW')
raw_novpn  = extract_block(html, 'RAW_NOVPN')

# ── Parser righe ─────────────────────────────────────────────────────────────
# Formato atteso (separatore: TAB reale o \t letterale):
#   NomeCinema - Città - NomeSala - NomeDispositivo[TAB]IP:porta[TAB]PROTOCOLLO
# Righe Coord / GEO → ignora; righe commento // → ignora; righe vuote → ignora

SKIP_KEYWORDS = ('Coord', 'GEO')

def parse_block(block, tipo):
    """
    Restituisce lista di dict con chiavi:
      cinema, citta, sala, dispositivo, ip, tipo
    """
    rows = []
    for raw_line in block.splitlines():
        line = raw_line.replace('\\t', '\t')
        line = line.strip()
        if not line:
            continue
        if line.startswith('//'):
            continue
        # Controlla se è una riga Coord/GEO (il terzo token dopo split su \t è GEO,
        # oppure la riga contiene letteralmente 'Coord' e termina con 'GEO')
        if 'Coord' in line and 'GEO' in line:
            continue
        # Split su TAB; accetta anche spazi multipli come separatori solo se non c'è tab
        parts = line.split('\t')
        if len(parts) < 2:
            # prova split su 2+ spazi
            parts = re.split(r'  +', line)
        if len(parts) < 2:
            continue

        left  = parts[0].strip()
        ip_porta = parts[1].strip() if len(parts) > 1 else ''
        # protocollo = parts[2] se esiste (non usato per Excel)

        # left = "Cinema - Città - Sala - Dispositivo"
        # Split su " - " (con spazi) — massimo 4 parti
        tokens = [t.strip() for t in left.split(' - ')]
        if len(tokens) < 4:
            # Potrebbe essere rete (es. "... - Rete - MikroTik") → ignoriamo
            # oppure una riga incompleta → ignoriamo
            continue

        cinema     = tokens[0]
        citta      = tokens[1]
        sala       = tokens[2]
        dispositivo = ' - '.join(tokens[3:])  # nel caso ci siano ulteriori " - "

        # Rimuovi la porta dall'IP
        ip = ip_porta.split(':')[0] if ':' in ip_porta else ip_porta

        rows.append({
            'cinema':      cinema,
            'citta':       citta,
            'sala':        sala,
            'dispositivo': dispositivo,
            'ip':          ip,
            'tipo':        tipo,
        })
    return rows

rows_vpn   = parse_block(raw_vpn,   'VPN')
rows_novpn = parse_block(raw_novpn, 'Offline')
all_rows   = rows_vpn + rows_novpn

print(f"Righe VPN:    {len(rows_vpn)}")
print(f"Righe Offline:{len(rows_novpn)}")
print(f"Totale dati:  {len(all_rows)}")

# ── Crea Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Dispositivi'

# Colonne e larghezze
columns = [
    ('Cinema',      32),
    ('Città',       22),
    ('Tipo',        10),
    ('Sala',        16),
    ('Dispositivo', 32),
    ('IP',          22),
]
for col_idx, (header, width) in enumerate(columns, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Intestazione (riga 1) ─────────────────────────────────────────────────────
header_fill = PatternFill(fill_type='solid', fgColor='1F3864')
header_font = Font(color='FFFFFF', bold=True)

for col_idx, (header, _) in enumerate(columns, start=1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill  = header_fill
    cell.font  = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

# ── Colori righe dati ─────────────────────────────────────────────────────────
FILL_PROIETTORE = PatternFill(fill_type='solid', fgColor='FFB3B3')
FILL_SERVER     = PatternFill(fill_type='solid', fgColor='FFFF99')
FILL_AUDIO      = PatternFill(fill_type='solid', fgColor='FFD580')

def get_fill(dispositivo):
    d = dispositivo.lower()
    if 'proiettore' in d:
        return FILL_PROIETTORE
    if 'server' in d:
        return FILL_SERVER
    if 'processore audio' in d:
        return FILL_AUDIO
    return None

# ── Scrivi righe ──────────────────────────────────────────────────────────────
for row_idx, row in enumerate(all_rows, start=2):
    values = [
        row['cinema'],
        row['citta'],
        row['tipo'],
        row['sala'],
        row['dispositivo'],
        row['ip'],
    ]
    fill = get_fill(row['dispositivo'])
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        if fill:
            cell.fill = fill

# ── Salva ─────────────────────────────────────────────────────────────────────
output_path = '/home/user/NOC-DASHBOARD-ROCCO/dispositivi_mancanti.xlsx'
wb.save(output_path)
print(f"\nSalvato: {output_path}")
print(f"Righe scritte (dati): {len(all_rows)}  (+ 1 intestazione = {len(all_rows)+1} totali)")
